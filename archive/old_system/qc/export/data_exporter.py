#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Data Exporter for Qualitative Coding Analysis
Comprehensive CSV and Excel export capabilities for analysis results
"""

import csv
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
from collections import defaultdict
import io

logger = logging.getLogger(__name__)

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    logger.warning("pandas not available - Excel export will be limited")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available - Excel formatting will be basic")

try:
    from docx import Document
    from docx.shared import Inches
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False
    logger.warning("python-docx not available - Word export will be unavailable")

try:
    import networkx as nx
    NETWORKX_AVAILABLE = True
except ImportError:
    NETWORKX_AVAILABLE = False
    logger.warning("networkx not available - GraphML export will be unavailable")


class DataExporter:
    """Export qualitative coding analysis data to CSV and Excel formats"""
    
    def __init__(self, output_dir: Union[str, Path] = "data/exports"):
        """
        Initialize data exporter
        
        Args:
            output_dir: Directory to save export files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
    def export_quotes_csv(self, interviews: List[Dict[str, Any]], output_file: str = None) -> str:
        """
        Export all quotes with codes to CSV format
        
        Args:
            interviews: List of interview analysis results
            output_file: Output filename (auto-generated if None)
            
        Returns:
            Path to generated CSV file
        """
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"quotes_export_{timestamp}.csv"
            
        output_path = self.output_dir / output_file
        
        # Prepare quote data - determine all possible fields first
        all_fields = set()
        quotes_data = []
        
        # First pass - collect all possible fields
        for interview in interviews:
            for quote in interview.get('quotes', []):
                if 'thematic_connection' in quote:
                    all_fields.update(['thematic_connection', 'connection_target', 
                                     'connection_confidence', 'connection_evidence'])
                if 'dialogue_turn' in quote:
                    all_fields.update(['turn_sequence', 'timestamp', 'turn_type'])
        
        # Second pass - create standardized rows
        for interview in interviews:
            interview_id = interview.get('interview_id', 'Unknown')
            
            for quote in interview.get('quotes', []):
                # Basic quote info
                row = {
                    'quote_id': quote.get('id', ''),
                    'interview_id': interview_id,
                    'speaker_name': quote.get('speaker', {}).get('name', ''),
                    'text': quote.get('text', ''),
                    'location_start': quote.get('location_start', ''),
                    'location_end': quote.get('location_end', ''),
                    'location_type': quote.get('location_type', ''),
                    'context_summary': quote.get('context_summary', ''),
                }
                
                # Add codes (multiple codes in separate columns or concatenated)
                code_ids = quote.get('code_ids', [])
                code_names = quote.get('code_names', [])
                code_confidences = quote.get('code_confidence_scores', [])
                
                row['code_ids'] = '|'.join(code_ids) if code_ids else ''
                row['code_names'] = '|'.join(code_names) if code_names else ''
                row['code_confidences'] = '|'.join([str(c) for c in code_confidences]) if code_confidences else ''
                
                # Add thematic connections (focus groups) - always include fields
                row['thematic_connection'] = quote.get('thematic_connection', '') if 'thematic_connection' in all_fields else ''
                row['connection_target'] = quote.get('connection_target', '') if 'connection_target' in all_fields else ''
                row['connection_confidence'] = quote.get('connection_confidence', '') if 'connection_confidence' in all_fields else ''
                row['connection_evidence'] = quote.get('connection_evidence', '') if 'connection_evidence' in all_fields else ''
                
                # Add dialogue turn info (focus groups) - always include fields  
                if 'dialogue_turn' in quote:
                    turn = quote['dialogue_turn']
                    row['turn_sequence'] = turn.get('sequence_number', '') if 'turn_sequence' in all_fields else ''
                    row['timestamp'] = turn.get('timestamp', '') if 'timestamp' in all_fields else ''
                    row['turn_type'] = turn.get('turn_type', '') if 'turn_type' in all_fields else ''
                else:
                    row['turn_sequence'] = '' if 'turn_sequence' in all_fields else None
                    row['timestamp'] = '' if 'timestamp' in all_fields else None
                    row['turn_type'] = '' if 'turn_type' in all_fields else None
                
                # Remove None values from row
                row = {k: v for k, v in row.items() if v is not None}
                
                quotes_data.append(row)
        
        # Write CSV
        if quotes_data:
            # Get all possible fieldnames across all rows
            all_fieldnames = set()
            for row in quotes_data:
                all_fieldnames.update(row.keys())
            
            fieldnames = list(all_fieldnames)
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(quotes_data)
        
        logger.info(f"Exported {len(quotes_data)} quotes to CSV: {output_path}")
        return str(output_path)
    
    def export_codes_csv(self, taxonomy: Dict[str, Any], output_file: str = None) -> str:
        """
        Export thematic codes taxonomy to CSV
        
        Args:
            taxonomy: Taxonomy data with codes
            output_file: Output filename (auto-generated if None)
            
        Returns:
            Path to generated CSV file
        """
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"codes_export_{timestamp}.csv"
            
        output_path = self.output_dir / output_file
        
        # Prepare codes data
        codes_data = []
        for code in taxonomy.get('codes', []):
            row = {
                'code_id': code.get('id', ''),
                'code_name': code.get('name', ''),
                'description': code.get('description', ''),
                'semantic_definition': code.get('semantic_definition', ''),
                'parent_id': code.get('parent_id', ''),
                'level': code.get('level', 0),
                'discovery_confidence': code.get('discovery_confidence', ''),
            }
            
            # Add example quotes (concatenated)
            examples = code.get('example_quotes', [])
            row['example_quotes'] = ' | '.join(examples) if examples else ''
            row['example_count'] = len(examples)
            
            codes_data.append(row)
        
        # Write CSV
        if codes_data:
            fieldnames = codes_data[0].keys()
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(codes_data)
        
        logger.info(f"Exported {len(codes_data)} codes to CSV: {output_path}")
        return str(output_path)
        
    def export_r_compatible_csv(self, interviews: List[Dict[str, Any]], output_file: str = None) -> str:
        """
        Export data in R-compatible format with standardized column names and types
        
        Args:
            interviews: List of interview data dictionaries
            output_file: Optional custom filename
            
        Returns:
            Path to generated R-compatible CSV file
        """
        if not output_file:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = f"qca_data_r_compatible_{timestamp}.csv"
        
        output_path = self.output_dir / output_file
        
        # Prepare R-compatible data structure
        r_data = []
        
        for interview in interviews:
            interview_id = interview.get('interview_id', 'Unknown')
            for quote in interview.get('quotes', []):
                # Create a flat structure suitable for R data.frame
                row = {
                    'quote_id': str(quote.get('id', '')),
                    'interview_id': str(interview_id),
                    'speaker_name': str(quote.get('speaker', {}).get('name', '')),
                    'quote_text': str(quote.get('text', '')),
                    'location_start': int(quote.get('location_start', 0) or 0),
                    'location_end': int(quote.get('location_end', 0) or 0),
                    'code_count': len(quote.get('code_names', [])),
                    'codes_list': '|'.join(quote.get('code_names', [])),
                }
                
                # Add binary indicators for each code (for QCA analysis)
                all_codes = set()
                for i in interviews:
                    for q in i.get('quotes', []):
                        all_codes.update(q.get('code_names', []))
                
                for code in sorted(all_codes):
                    # Create R-compatible variable names (no spaces, special chars)
                    r_var_name = f"code_{code.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '')}"
                    row[r_var_name] = 1 if code in quote.get('code_names', []) else 0
                
                r_data.append(row)
        
        # Write R-compatible CSV with proper headers
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            if r_data:
                writer = csv.DictWriter(csvfile, fieldnames=r_data[0].keys())
                writer.writeheader()
                writer.writerows(r_data)
        
        logger.info(f"Exported R-compatible data with {len(r_data)} rows to: {output_path}")
        return str(output_path)
        
    def generate_r_import_script(self, csv_file_path: str, output_file: str = None) -> str:
        """
        Generate R script to import QCA data and perform basic analysis
        
        Args:
            csv_file_path: Path to the R-compatible CSV file
            output_file: Optional custom filename for R script
            
        Returns:
            Path to generated R script
        """
        if not output_file:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = f"qca_analysis_import_{timestamp}.R"
        
        output_path = self.output_dir / output_file
        csv_filename = Path(csv_file_path).name
        
        r_script_content = f'''#!/usr/bin/env Rscript
# QCA Analysis Import Script
# Generated by qualitative_coding pipeline on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
# 
# This script loads QCA data and performs basic qualitative comparative analysis

# Required packages
required_packages <- c("QCA", "dplyr", "readr", "tidyr")
for (pkg in required_packages) {{
    if (!requireNamespace(pkg, quietly = TRUE)) {{
        cat("Installing package:", pkg, "\\n")
        install.packages(pkg)
    }}
    library(pkg, character.only = TRUE)
}}

# Load data
cat("Loading QCA data from {csv_filename}...\\n")
qca_data <- read_csv("{csv_filename}", show_col_types = FALSE)

# Display basic information
cat("\\n=== Data Summary ===\\n")
cat("Total observations:", nrow(qca_data), "\\n")
cat("Total variables:", ncol(qca_data), "\\n")
cat("Interviews:", length(unique(qca_data$interview_id)), "\\n")
cat("Speakers:", length(unique(qca_data$speaker_name)), "\\n")

# Identify code columns (binary indicators)
code_cols <- grep("^code_", names(qca_data), value = TRUE)
cat("\\nCode variables found:", length(code_cols), "\\n")
print(code_cols)

# Basic frequency analysis
cat("\\n=== Code Frequencies ===\\n")
code_freq <- qca_data[code_cols] %>%
    summarise_all(sum) %>%
    gather(code, frequency) %>%
    arrange(desc(frequency))
print(code_freq)

# Prepare for QCA analysis (if you have outcome variables)
# Example: Create a simplified truth table
if (length(code_cols) >= 3 && length(code_cols) <= 10) {{
    cat("\\n=== QCA Analysis Preparation ===\\n")
    
    # Select top codes for analysis (limit to manageable number)
    top_codes <- code_freq$code[1:min(5, length(code_cols))]
    
    # Create analysis dataset
    qca_analysis_data <- qca_data[c("interview_id", top_codes)]
    
    # Aggregate by interview (sum codes per interview)
    qca_summary <- qca_analysis_data %>%
        group_by(interview_id) %>%
        summarise_all(~ as.numeric(sum(.) > 0), .groups = "drop")
    
    cat("Aggregated data for", nrow(qca_summary), "interviews\\n")
    print(head(qca_summary))
    
    # Example: Create truth table (you'll need to define your outcome variable)
    # Uncomment and modify as needed:
    # outcome_var <- "your_outcome_variable"
    # truth_table <- truthTable(qca_summary, outcome = outcome_var, 
    #                          conditions = top_codes[-length(top_codes)])
    # print(truth_table)
    
}} else {{
    cat("\\nSkipping QCA analysis - too many or too few code variables\\n")
    cat("For QCA analysis, typically need 3-10 condition variables\\n")
}}

# Export processed data for further analysis
processed_file <- sub("\\.csv$", "_processed.csv", "{csv_filename}")
write_csv(qca_data, processed_file)
cat("\\nProcessed data saved to:", processed_file, "\\n")

cat("\\n=== Analysis Complete ===\\n")
cat("Next steps:\\n")
cat("1. Review code frequencies and select conditions for QCA\\n")
cat("2. Define outcome variable(s)\\n")
cat("3. Use QCA package functions: truthTable(), minimize(), pof()\\n")
cat("4. For more info: help(package='QCA')\\n")
'''
        
        # Write R script
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(r_script_content)
        
        logger.info(f"Generated R import script: {output_path}")
        return str(output_path)
        
    def export_r_package_integration(self, interviews: List[Dict[str, Any]], base_filename: str = None) -> Dict[str, str]:
        """
        Complete R package integration export - generates both CSV data and R script
        
        Args:
            interviews: List of interview data dictionaries
            base_filename: Base filename for generated files
            
        Returns:
            Dictionary with paths to generated files
        """
        if not base_filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            base_filename = f"qca_r_export_{timestamp}"
        
        # Generate R-compatible CSV
        csv_file = f"{base_filename}.csv"
        csv_path = self.export_r_compatible_csv(interviews, csv_file)
        
        # Generate R import script
        r_script_file = f"{base_filename}_analysis.R"
        r_script_path = self.generate_r_import_script(csv_path, r_script_file)
        
        # Generate additional documentation
        readme_file = f"{base_filename}_README.md"
        readme_path = self.generate_r_integration_readme(csv_path, r_script_path, readme_file)
        
        logger.info(f"R package integration export complete - 3 files generated")
        return {
            'csv_data': csv_path,
            'r_script': r_script_path,
            'readme': readme_path
        }
        
    def generate_r_integration_readme(self, csv_path: str, r_script_path: str, output_file: str) -> str:
        """
        Generate README documentation for R integration
        
        Args:
            csv_path: Path to CSV data file
            r_script_path: Path to R script file
            output_file: Output filename for README
            
        Returns:
            Path to generated README file
        """
        output_path = self.output_dir / output_file
        csv_filename = Path(csv_path).name
        r_script_filename = Path(r_script_path).name
        
        readme_content = f'''# QCA R Package Integration

Generated by qualitative_coding pipeline on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## Files Included

- **{csv_filename}**: R-compatible CSV data file
- **{r_script_filename}**: R analysis script
- **{output_file}**: This documentation file

## Quick Start

### 1. Prerequisites

Make sure you have R installed with the following packages:

```r
install.packages(c("QCA", "dplyr", "readr", "tidyr"))
```

### 2. Load and Analyze Data

```r
# Run the analysis script
source("{r_script_filename}")

# Or load data manually
library(readr)
qca_data <- read_csv("{csv_filename}")
```

### 3. Data Structure

The CSV file contains the following columns:

- `quote_id`: Unique identifier for each quote
- `interview_id`: Interview identifier
- `speaker_name`: Name of the speaker
- `quote_text`: The actual quote text
- `location_start`, `location_end`: Position in source document
- `code_count`: Number of codes applied to this quote
- `codes_list`: Pipe-separated list of code names
- `code_*`: Binary indicators (0/1) for each thematic code

## Common QCA Analysis Workflows

### Basic Frequency Analysis

```r
# Load libraries
library(QCA)
library(dplyr)

# Load data
data <- read_csv("{csv_filename}")

# Get code columns
code_cols <- grep("^code_", names(data), value = TRUE)

# Frequency analysis
code_freq <- data[code_cols] %>%
    summarise_all(sum) %>%
    gather(code, frequency) %>%
    arrange(desc(frequency))
```

### Truth Table Analysis

```r
# Aggregate by interview
interview_summary <- data %>%
    group_by(interview_id) %>%
    summarise_at(vars(starts_with("code_")), ~ as.numeric(sum(.) > 0))

# Create truth table (define your outcome variable)
# outcome_var <- "your_outcome_variable"  # Choose your outcome
# conditions <- c("code_condition1", "code_condition2", "code_condition3")
# tt <- truthTable(interview_summary, outcome = outcome_var, conditions = conditions)
```

### QCA Minimization

```r
# After creating truth table
# conservative <- minimize(tt, details = TRUE)
# parsimonious <- minimize(tt, include = "?", details = TRUE)
```

## Integration with Other R Packages

### Text Analysis
```r
library(tm)
library(quanteda)

# Analyze quote texts
corpus <- corpus(data$quote_text)
```

### Network Analysis
```r
library(igraph)

# Create code co-occurrence networks
# (implementation depends on your analysis needs)
```

### Visualization
```r
library(ggplot2)
library(plotly)

# Visualize code frequencies
ggplot(code_freq, aes(x = reorder(code, frequency), y = frequency)) +
    geom_bar(stat = "identity") +
    coord_flip() +
    labs(title = "Code Frequency Analysis")
```

## Support

- QCA Package Documentation: `help(package="QCA")`
- Qualitative Coding Analysis Tool: [Project Documentation]
- For issues with the R integration, check the original data export logs

## Generated Files

This export was created from:
- CSV Data File: {csv_filename}  
- R Script: {r_script_filename}
- Export Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

---

*This R integration export is designed for academic research workflows and is compatible with standard QCA analysis packages.*
'''
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(readme_content)
            
        logger.info(f"Generated R integration README: {output_path}")
        return str(output_path)
    
    def export_speakers_csv(self, speaker_schema: Dict[str, Any], interviews: List[Dict[str, Any]], output_file: str = None) -> str:
        """
        Export speaker information to CSV
        
        Args:
            speaker_schema: Speaker schema definition
            interviews: Interview data with speaker information
            output_file: Output filename (auto-generated if None)
            
        Returns:
            Path to generated CSV file
        """
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"speakers_export_{timestamp}.csv"
            
        output_path = self.output_dir / output_file
        
        # Collect unique speakers
        speakers_map = {}
        
        for interview in interviews:
            interview_id = interview.get('interview_id', 'Unknown')
            
            # From speaker objects in quotes
            for quote in interview.get('quotes', []):
                speaker = quote.get('speaker', {})
                if speaker and speaker.get('name'):
                    speaker_name = speaker['name']
                    if speaker_name not in speakers_map:
                        speakers_map[speaker_name] = {
                            'name': speaker_name,
                            'interviews': [],
                            'quotes_count': 0,
                            'confidence': speaker.get('confidence', ''),
                            'identification_method': speaker.get('identification_method', ''),
                        }
                        
                        # Add speaker properties from schema
                        for prop in speaker_schema.get('properties', []):
                            prop_name = prop.get('property_name', '')
                            speakers_map[speaker_name][prop_name] = speaker.get(prop_name, '')
                    
                    # Track interview participation
                    if interview_id not in speakers_map[speaker_name]['interviews']:
                        speakers_map[speaker_name]['interviews'].append(interview_id)
                    speakers_map[speaker_name]['quotes_count'] += 1
        
        # Prepare speakers data
        speakers_data = []
        for speaker_name, speaker_data in speakers_map.items():
            row = speaker_data.copy()
            row['interviews'] = '|'.join(row['interviews'])
            speakers_data.append(row)
        
        # Write CSV
        if speakers_data:
            fieldnames = speakers_data[0].keys() if speakers_data else []
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(speakers_data)
        
        logger.info(f"Exported {len(speakers_data)} speakers to CSV: {output_path}")
        return str(output_path)
    
    def export_entities_csv(self, entity_schema: Dict[str, Any], interviews: List[Dict[str, Any]], output_file: str = None) -> str:
        """
        Export entities and relationships to CSV
        
        Args:
            entity_schema: Entity schema definition
            interviews: Interview data with entity information
            output_file: Output filename (auto-generated if None)
            
        Returns:
            Path to generated CSV file
        """
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"entities_export_{timestamp}.csv"
            
        output_path = self.output_dir / output_file
        
        # Collect entities from interviews
        entities_data = []
        
        for interview in interviews:
            interview_id = interview.get('interview_id', 'Unknown')
            
            # From entity extractions
            for entity in interview.get('entities', []):
                row = {
                    'entity_id': entity.get('id', ''),
                    'entity_name': entity.get('name', ''),
                    'entity_type': entity.get('entity_type', ''),
                    'interview_id': interview_id,
                    'confidence': entity.get('confidence', ''),
                    'context': entity.get('context', ''),
                    'line_range': entity.get('line_range', ''),
                }
                
                # Add relationships
                relationships = entity.get('relationships', [])
                if relationships:
                    row['relationships'] = ' | '.join([
                        f"{rel.get('target', '')} ({rel.get('relationship_type', '')})"
                        for rel in relationships
                    ])
                    row['relationship_count'] = len(relationships)
                else:
                    row['relationships'] = ''
                    row['relationship_count'] = 0
                
                entities_data.append(row)
        
        # Write CSV
        if entities_data:
            fieldnames = entities_data[0].keys() if entities_data else []
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(entities_data)
        
        logger.info(f"Exported {len(entities_data)} entities to CSV: {output_path}")
        return str(output_path)
    
    def export_complete_analysis_csv(self, analysis_data: Dict[str, Any], output_prefix: str = None) -> Dict[str, str]:
        """
        Export complete analysis to multiple CSV files
        
        Args:
            analysis_data: Complete analysis data including interviews, taxonomy, etc.
            output_prefix: Prefix for output files (auto-generated if None)
            
        Returns:
            Dictionary mapping export type to file path
        """
        if not output_prefix:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_prefix = f"complete_analysis_{timestamp}"
        
        export_paths = {}
        
        # Export quotes
        if 'interviews' in analysis_data:
            quotes_file = f"{output_prefix}_quotes.csv"
            export_paths['quotes'] = self.export_quotes_csv(
                analysis_data['interviews'], quotes_file
            )
        
        # Export codes
        if 'taxonomy' in analysis_data:
            codes_file = f"{output_prefix}_codes.csv"
            export_paths['codes'] = self.export_codes_csv(
                analysis_data['taxonomy'], codes_file
            )
        
        # Export speakers
        if 'speaker_schema' in analysis_data and 'interviews' in analysis_data:
            speakers_file = f"{output_prefix}_speakers.csv"
            export_paths['speakers'] = self.export_speakers_csv(
                analysis_data['speaker_schema'], analysis_data['interviews'], speakers_file
            )
        
        # Export entities
        if 'entity_schema' in analysis_data and 'interviews' in analysis_data:
            entities_file = f"{output_prefix}_entities.csv"
            export_paths['entities'] = self.export_entities_csv(
                analysis_data['entity_schema'], analysis_data['interviews'], entities_file
            )
        
        logger.info(f"Exported complete analysis to {len(export_paths)} CSV files")
        return export_paths
    
    def export_excel_workbook(self, analysis_data: Dict[str, Any], output_file: str = None) -> str:
        """
        Export complete analysis to Excel workbook with multiple sheets
        
        Args:
            analysis_data: Complete analysis data
            output_file: Output filename (auto-generated if None)
            
        Returns:
            Path to generated Excel file
        """
        if not PANDAS_AVAILABLE and not OPENPYXL_AVAILABLE:
            raise ImportError("Either pandas or openpyxl required for Excel export")
        
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"analysis_export_{timestamp}.xlsx"
            
        output_path = self.output_dir / output_file
        
        if PANDAS_AVAILABLE:
            return self._export_excel_pandas(analysis_data, output_path)
        else:
            return self._export_excel_openpyxl(analysis_data, output_path)
    
    def _export_excel_pandas(self, analysis_data: Dict[str, Any], output_path: Path) -> str:
        """Export using pandas (preferred method)"""
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            
            # Quotes sheet
            if 'interviews' in analysis_data:
                quotes_data = self._prepare_quotes_dataframe(analysis_data['interviews'])
                if not quotes_data.empty:
                    quotes_data.to_excel(writer, sheet_name='Quotes', index=False)
            
            # Codes sheet
            if 'taxonomy' in analysis_data:
                codes_data = self._prepare_codes_dataframe(analysis_data['taxonomy'])
                if not codes_data.empty:
                    codes_data.to_excel(writer, sheet_name='Codes', index=False)
            
            # Speakers sheet
            if 'speaker_schema' in analysis_data and 'interviews' in analysis_data:
                speakers_data = self._prepare_speakers_dataframe(
                    analysis_data['speaker_schema'], analysis_data['interviews']
                )
                if not speakers_data.empty:
                    speakers_data.to_excel(writer, sheet_name='Speakers', index=False)
            
            # Entities sheet
            if 'entity_schema' in analysis_data and 'interviews' in analysis_data:
                entities_data = self._prepare_entities_dataframe(
                    analysis_data['entity_schema'], analysis_data['interviews']
                )
                if not entities_data.empty:
                    entities_data.to_excel(writer, sheet_name='Entities', index=False)
            
            # Summary sheet
            summary_data = self._prepare_summary_dataframe(analysis_data)
            summary_data.to_excel(writer, sheet_name='Summary', index=False)
        
        logger.info(f"Exported complete analysis to Excel: {output_path}")
        return str(output_path)
    
    def _export_excel_openpyxl(self, analysis_data: Dict[str, Any], output_path: Path) -> str:
        """Export using openpyxl (fallback method)"""
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create sheets with data
        self._create_excel_sheet_openpyxl(wb, 'Summary', self._prepare_summary_data(analysis_data))
        
        if 'interviews' in analysis_data:
            quotes_data = self._prepare_quotes_data(analysis_data['interviews'])
            if quotes_data:
                self._create_excel_sheet_openpyxl(wb, 'Quotes', quotes_data)
        
        if 'taxonomy' in analysis_data:
            codes_data = self._prepare_codes_data(analysis_data['taxonomy'])
            if codes_data:
                self._create_excel_sheet_openpyxl(wb, 'Codes', codes_data)
        
        wb.save(output_path)
        logger.info(f"Exported complete analysis to Excel: {output_path}")
        return str(output_path)
    
    def _prepare_quotes_dataframe(self, interviews: List[Dict[str, Any]]) -> 'pd.DataFrame':
        """Prepare quotes data as pandas DataFrame"""
        quotes_data = []
        for interview in interviews:
            interview_id = interview.get('interview_id', 'Unknown')
            for quote in interview.get('quotes', []):
                row = {
                    'Quote ID': quote.get('id', ''),
                    'Interview ID': interview_id,
                    'Speaker': quote.get('speaker', {}).get('name', ''),
                    'Text': quote.get('text', ''),
                    'Codes': '|'.join(quote.get('code_names', [])),
                    'Location': f"{quote.get('location_start', '')}-{quote.get('location_end', '')}",
                    'Context': quote.get('context_summary', ''),
                }
                
                # Focus group specific fields
                if 'thematic_connection' in quote:
                    row['Thematic Connection'] = quote.get('thematic_connection', '')
                    row['Connection Target'] = quote.get('connection_target', '')
                
                quotes_data.append(row)
        
        return pd.DataFrame(quotes_data) if quotes_data else pd.DataFrame()
    
    def _prepare_codes_dataframe(self, taxonomy: Dict[str, Any]) -> 'pd.DataFrame':
        """Prepare codes data as pandas DataFrame"""
        codes_data = []
        for code in taxonomy.get('codes', []):
            codes_data.append({
                'Code ID': code.get('id', ''),
                'Name': code.get('name', ''),
                'Description': code.get('description', ''),
                'Level': code.get('level', 0),
                'Parent ID': code.get('parent_id', ''),
                'Example Count': len(code.get('example_quotes', [])),
                'Discovery Confidence': code.get('discovery_confidence', ''),
            })
        
        return pd.DataFrame(codes_data) if codes_data else pd.DataFrame()
    
    def _prepare_speakers_dataframe(self, speaker_schema: Dict[str, Any], interviews: List[Dict[str, Any]]) -> 'pd.DataFrame':
        """Prepare speakers data as pandas DataFrame"""
        speakers_map = {}
        
        for interview in interviews:
            for quote in interview.get('quotes', []):
                speaker = quote.get('speaker', {})
                if speaker and speaker.get('name'):
                    name = speaker['name']
                    if name not in speakers_map:
                        speakers_map[name] = {
                            'Name': name,
                            'Quotes Count': 0,
                            'Confidence': speaker.get('confidence', ''),
                            'Identification Method': speaker.get('identification_method', ''),
                        }
                    speakers_map[name]['Quotes Count'] += 1
        
        speakers_data = list(speakers_map.values())
        return pd.DataFrame(speakers_data) if speakers_data else pd.DataFrame()
    
    def _prepare_entities_dataframe(self, entity_schema: Dict[str, Any], interviews: List[Dict[str, Any]]) -> 'pd.DataFrame':
        """Prepare entities data as pandas DataFrame"""
        entities_data = []
        
        for interview in interviews:
            for entity in interview.get('entities', []):
                entities_data.append({
                    'Entity ID': entity.get('id', ''),
                    'Name': entity.get('name', ''),
                    'Type': entity.get('entity_type', ''),
                    'Interview': interview.get('interview_id', ''),
                    'Confidence': entity.get('confidence', ''),
                    'Relationships': len(entity.get('relationships', [])),
                })
        
        return pd.DataFrame(entities_data) if entities_data else pd.DataFrame()
    
    def _prepare_summary_dataframe(self, analysis_data: Dict[str, Any]) -> 'pd.DataFrame':
        """Prepare summary statistics as pandas DataFrame"""
        interviews = analysis_data.get('interviews', [])
        taxonomy = analysis_data.get('taxonomy', {})
        
        total_quotes = sum(len(interview.get('quotes', [])) for interview in interviews)
        total_codes = len(taxonomy.get('codes', []))
        total_entities = sum(len(interview.get('entities', [])) for interview in interviews)
        
        summary_data = [
            {'Metric': 'Total Interviews', 'Value': len(interviews)},
            {'Metric': 'Total Quotes', 'Value': total_quotes},
            {'Metric': 'Total Codes', 'Value': total_codes},
            {'Metric': 'Total Entities', 'Value': total_entities},
            {'Metric': 'Generated At', 'Value': datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
        ]
        
        return pd.DataFrame(summary_data)
    
    def _create_excel_sheet_openpyxl(self, workbook: 'Workbook', sheet_name: str, data: List[Dict[str, Any]]) -> None:
        """Create Excel sheet using openpyxl"""
        if not data:
            return
            
        ws = workbook.create_sheet(title=sheet_name)
        
        # Headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        for row, item in enumerate(data, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=item.get(header, ''))
        
        # Adjust column widths
        for column in ws.columns:
            length = max(len(str(cell.value or '')) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(length + 2, 50)
    
    def _prepare_quotes_data(self, interviews: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Prepare quotes data for openpyxl"""
        quotes_data = []
        for interview in interviews:
            interview_id = interview.get('interview_id', 'Unknown')
            for quote in interview.get('quotes', []):
                quotes_data.append({
                    'Quote ID': quote.get('id', ''),
                    'Interview ID': interview_id,
                    'Speaker': quote.get('speaker', {}).get('name', ''),
                    'Text': quote.get('text', ''),
                    'Codes': '|'.join(quote.get('code_names', [])),
                    'Location': f"{quote.get('location_start', '')}-{quote.get('location_end', '')}",
                })
        return quotes_data
    
    def _prepare_codes_data(self, taxonomy: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Prepare codes data for openpyxl"""
        return [{
            'Code ID': code.get('id', ''),
            'Name': code.get('name', ''),
            'Description': code.get('description', ''),
            'Level': code.get('level', 0),
        } for code in taxonomy.get('codes', [])]
    
    def _prepare_summary_data(self, analysis_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Prepare summary data for openpyxl"""
        interviews = analysis_data.get('interviews', [])
        taxonomy = analysis_data.get('taxonomy', {})
        
        return [
            {'Metric': 'Total Interviews', 'Value': len(interviews)},
            {'Metric': 'Total Quotes', 'Value': sum(len(i.get('quotes', [])) for i in interviews)},
            {'Metric': 'Total Codes', 'Value': len(taxonomy.get('codes', []))},
            {'Metric': 'Generated At', 'Value': datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
        ]
    
    def export_latex_report(self, analysis_data: Dict[str, Any], output_file: str = None) -> str:
        """
        Export analysis results as LaTeX document for academic publications
        
        Args:
            analysis_data: Complete analysis data
            output_file: Output filename (auto-generated if None)
            
        Returns:
            Path to generated LaTeX file
        """
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"qca_analysis_report_{timestamp}.tex"
            
        output_path = self.output_dir / output_file
        
        interviews = analysis_data.get('interviews', [])
        taxonomy = analysis_data.get('taxonomy', {})
        
        # Calculate summary statistics
        total_quotes = sum(len(interview.get('quotes', [])) for interview in interviews)
        total_codes = len(taxonomy.get('codes', []))
        
        # Build LaTeX content
        latex_content = f'''\\documentclass{{article}}
\\usepackage[utf8]{{inputenc}}
\\usepackage{{longtable}}
\\usepackage{{booktabs}}
\\usepackage{{geometry}}
\\usepackage{{hyperref}}
\\geometry{{margin=1in}}

\\title{{Qualitative Coding Analysis Report}}
\\author{{Generated by QCA Pipeline}}
\\date{{{datetime.now().strftime('%B %d, %Y')}}}

\\begin{{document}}

\\maketitle

\\section{{Executive Summary}}

This report presents the results of a qualitative coding analysis conducted using the QCA Pipeline system. The analysis encompasses {len(interviews)} interviews with a total of {total_quotes} coded quotes across {total_codes} thematic codes.

\\section{{Analysis Overview}}

\\begin{{itemize}}
    \\item Total Interviews: {len(interviews)}
    \\item Total Quotes: {total_quotes}
    \\item Total Codes: {total_codes}
    \\item Analysis Date: {datetime.now().strftime('%Y-%m-%d')}
\\end{{itemize}}

\\section{{Thematic Code Structure}}

'''
        
        # Add codes table
        if taxonomy.get('codes'):
            latex_content += '''\\begin{longtable}{p{2cm}p{4cm}p{6cm}p{1cm}}
\\toprule
Code ID & Name & Description & Level \\\\
\\midrule
\\endhead
'''
            
            for code in taxonomy.get('codes', []):
                code_id = self._latex_escape(code.get('id', ''))
                name = self._latex_escape(code.get('name', ''))
                desc = self._latex_escape(code.get('description', '')[:100])
                level = code.get('level', 0)
                
                latex_content += f"{code_id} & {name} & {desc} & {level} \\\\\n"
            
            latex_content += '''\\bottomrule
\\end{longtable}

'''
        
        # Add quote samples
        latex_content += '''\\section{Quote Analysis}

\\subsection{Sample Coded Quotes}

'''
        
        sample_count = 0
        for interview in interviews[:3]:  # Sample from first 3 interviews
            interview_id = self._latex_escape(interview.get('interview_id', 'Unknown'))
            latex_content += f"\\subsubsection{{Interview: {interview_id}}}\n\n"
            
            for quote in interview.get('quotes', [])[:5]:  # Sample 5 quotes per interview
                if sample_count >= 15:  # Limit total samples
                    break
                    
                text = self._latex_escape(quote.get('text', '')[:200])
                codes = ', '.join([self._latex_escape(c) for c in quote.get('code_names', [])])
                speaker = self._latex_escape(quote.get('speaker', {}).get('name', 'Unknown'))
                
                latex_content += f'''\\begin{{quote}}
    \\"{text}...\\" \\\\
    \\textit{{Speaker: {speaker}}} \\\\
    \\textbf{{Codes:}} {codes}
\\end{{quote}}

'''
                sample_count += 1
            
            if sample_count >= 15:
                break
        
        # Add methodology section
        latex_content += '''\\section{Methodology}

\\subsection{Data Collection}
Qualitative interviews were analyzed using systematic thematic coding approaches with LLM-assisted pattern recognition.

\\subsection{Coding Process}
Codes were generated through both inductive and deductive approaches, with confidence scoring and validation.

\\subsection{Analysis Framework}
The analysis employed qualitative comparative analysis (QCA) principles for systematic pattern identification.

\\section{Conclusions}

This analysis provides a comprehensive view of the qualitative data through systematic coding and thematic analysis. The identified patterns suggest significant themes that warrant further investigation.

\\appendix
\\section{Technical Details}

\\begin{itemize}
    \\item Analysis System: QCA Pipeline
    \\item Generated: ''' + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + '''
    \\item Export Format: LaTeX Academic Report
\\end{itemize}

\\end{document}
'''
        
        # Write LaTeX file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(latex_content)
        
        logger.info(f"Exported LaTeX report: {output_path}")
        return str(output_path)
    
    def _latex_escape(self, text: str) -> str:
        """
        Escape special LaTeX characters
        
        Args:
            text: Input text to escape
            
        Returns:
            LaTeX-safe text
        """
        if not text:
            return ''
        
        # LaTeX special characters that need escaping
        escape_chars = {
            '&': '\\&',
            '%': '\\%',
            '$': '\\$',
            '#': '\\#',
            '^': '\\textasciicircum{}',
            '_': '\\_',
            '{': '\\{',
            '}': '\\}',
            '~': '\\textasciitilde{}',
            '\\': '\\textbackslash{}',
        }
        
        result = str(text)
        for char, escape in escape_chars.items():
            result = result.replace(char, escape)
        
        return result
    
    def export_spss_syntax(self, interviews: List[Dict[str, Any]], output_file: str = None) -> str:
        """
        Export SPSS syntax file for importing and analyzing QCA data
        
        Args:
            interviews: List of interview data dictionaries
            output_file: Output filename (auto-generated if None)
            
        Returns:
            Path to generated SPSS syntax file
        """
        from .academic_exporters import AcademicExporters
        
        academic_exporter = AcademicExporters(self.output_dir)
        return academic_exporter.export_spss_syntax(interviews, output_file)
    
    def export_word_report(self, analysis_data: Dict[str, Any], output_file: str = None) -> str:
        """
        Export analysis results as Microsoft Word document
        
        Args:
            analysis_data: Complete analysis data
            output_file: Output filename (auto-generated if None)
            
        Returns:
            Path to generated Word file
        """
        from .academic_exporters import AcademicExporters
        
        academic_exporter = AcademicExporters(self.output_dir)
        return academic_exporter.export_word_report(analysis_data, output_file)
    
    def export_graphml_network(self, interviews: List[Dict[str, Any]], output_file: str = None) -> str:
        """
        Export code co-occurrence network as GraphML format for network analysis
        
        Args:
            interviews: List of interview data dictionaries
            output_file: Output filename (auto-generated if None)
            
        Returns:
            Path to generated GraphML file
        """
        from .academic_exporters import AcademicExporters
        
        academic_exporter = AcademicExporters(self.output_dir)
        return academic_exporter.export_graphml_network(interviews, output_file)
    
    def export_academic_complete(self, analysis_data: Dict[str, Any], base_filename: str = None) -> Dict[str, str]:
        """
        Complete academic export package - generates all academic formats
        
        Args:
            analysis_data: Complete analysis data
            base_filename: Base filename for generated files
            
        Returns:
            Dictionary with paths to all generated files
        """
        if not base_filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            base_filename = f"qca_academic_export_{timestamp}"
        
        export_files = {}
        interviews = analysis_data.get('interviews', [])
        
        try:
            # LaTeX report
            latex_file = f"{base_filename}_report.tex"
            export_files['latex_report'] = self.export_latex_report(analysis_data, latex_file)
        except Exception as e:
            logger.error(f"LaTeX export failed: {e}")
            export_files['latex_error'] = str(e)
        
        try:
            # SPSS syntax
            spss_file = f"{base_filename}_import.sps"
            export_files['spss_syntax'] = self.export_spss_syntax(interviews, spss_file)
        except Exception as e:
            logger.error(f"SPSS export failed: {e}")
            export_files['spss_error'] = str(e)
        
        try:
            # Word report
            if DOCX_AVAILABLE:
                word_file = f"{base_filename}_report.docx"
                export_files['word_report'] = self.export_word_report(analysis_data, word_file)
            else:
                export_files['word_unavailable'] = "python-docx not installed"
        except Exception as e:
            logger.error(f"Word export failed: {e}")
            export_files['word_error'] = str(e)
        
        try:
            # GraphML network
            if NETWORKX_AVAILABLE:
                graphml_file = f"{base_filename}_network.graphml"
                export_files['graphml_network'] = self.export_graphml_network(interviews, graphml_file)
            else:
                export_files['graphml_unavailable'] = "networkx not installed"
        except Exception as e:
            logger.error(f"GraphML export failed: {e}")
            export_files['graphml_error'] = str(e)
        
        # R package integration (already existed)
        try:
            r_files = self.export_r_package_integration(interviews, base_filename)
            export_files.update({f'r_{k}': v for k, v in r_files.items()})
        except Exception as e:
            logger.error(f"R export failed: {e}")
            export_files['r_error'] = str(e)
        
        # Excel workbook (already existed) 
        try:
            excel_file = f"{base_filename}_data.xlsx"
            export_files['excel_workbook'] = self.export_excel_workbook(analysis_data, excel_file)
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            export_files['excel_error'] = str(e)
        
        # Generate comprehensive README
        try:
            readme_file = f"{base_filename}_README.md"
            export_files['readme'] = self._generate_academic_readme(export_files, readme_file)
        except Exception as e:
            logger.error(f"README generation failed: {e}")
            export_files['readme_error'] = str(e)
        
        successful_exports = [k for k in export_files.keys() if not k.endswith('_error') and not k.endswith('_unavailable')]
        logger.info(f"Academic export complete - {len(successful_exports)} formats generated")
        return export_files
    
    def _generate_academic_readme(self, export_files: Dict[str, str], output_file: str) -> str:
        """
        Generate comprehensive README for academic export package
        
        Args:
            export_files: Dictionary of exported file paths
            output_file: Output filename for README
            
        Returns:
            Path to generated README file
        """
        output_path = self.output_dir / output_file
        
        readme_content = f'''# QCA Academic Export Package

Generated by QCA Pipeline on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## Package Contents

This academic export package contains analysis results in multiple formats suitable for academic research and publication.

### Available Formats

'''
        
        # Document available files
        format_descriptions = {
            'latex_report': 'LaTeX academic report ready for compilation and publication',
            'word_report': 'Microsoft Word report for collaborative editing and review',
            'spss_syntax': 'SPSS syntax file for statistical analysis and data import',
            'graphml_network': 'GraphML network file for code co-occurrence analysis',
            'excel_workbook': 'Excel workbook with multiple data sheets',
            'r_csv': 'R-compatible CSV data file',
            'r_script': 'R analysis script with QCA examples',
            'r_readme': 'R integration documentation'
        }
        
        for format_key, description in format_descriptions.items():
            if format_key in export_files:
                filename = Path(export_files[format_key]).name
                readme_content += f"- **{filename}**: {description}\n"
        
        # Document unavailable formats
        unavailable = [k for k in export_files.keys() if k.endswith('_unavailable')]
        if unavailable:
            readme_content += "\n### Unavailable Formats\n\n"
            for key in unavailable:
                format_name = key.replace('_unavailable', '').title()
                reason = export_files[key]
                readme_content += f"- **{format_name}**: {reason}\n"
        
        # Usage instructions
        readme_content += '''

## Usage Instructions

### LaTeX Report
```bash
# Compile LaTeX document
pdflatex report.tex
```

### SPSS Analysis
1. Open SPSS
2. Run syntax file: `File > Open > Syntax`
3. Execute all commands or run selectively

### R Analysis
```r
# Load QCA data
source("analysis_script.R")

# Or load data manually
library(readr)
data <- read_csv("data.csv")
```

### Network Analysis
1. Open GraphML file in Gephi, Cytoscape, or NetworkX
2. Analyze code co-occurrence patterns
3. Visualize thematic relationships

### Excel Analysis
1. Open Excel workbook
2. Review data across multiple sheets
3. Create pivot tables and charts

## Academic Citation

When using this analysis in academic work, please consider citing the methodology and tools used:

```
Qualitative coding analysis performed using QCA Pipeline system with LLM-assisted 
thematic identification. Generated academic export package includes multiple formats 
for reproducible research.
```

## Software Requirements

### LaTeX
- LaTeX distribution (TeX Live, MiKTeX, or MacTeX)
- Required packages: longtable, booktabs, geometry, hyperref

### SPSS
- IBM SPSS Statistics (version 25 or higher recommended)
- Base module sufficient for basic analysis

### R
- R (version 4.0 or higher)
- Required packages: QCA, dplyr, readr, tidyr

### Network Analysis
- Gephi (recommended for visualization)
- Cytoscape (for biological-style networks) 
- Python NetworkX (for programmatic analysis)
- R igraph package

## Troubleshooting

### LaTeX Issues
- Missing packages: Install via package manager
- Encoding errors: Ensure UTF-8 encoding

### SPSS Issues
- Variable name limits: SPSS limits variable names to 64 characters
- Encoding: Use UTF-8 for international characters

### R Issues
- Package installation: `install.packages(c("QCA", "dplyr", "readr", "tidyr"))`
- Memory issues: Use data.table for large datasets

## Support

- QCA Pipeline Documentation: [Project Repository]
- Academic Export Issues: Check individual format documentation
- Integration Problems: Verify software versions and dependencies

---

*This export package supports reproducible qualitative research workflows across multiple analysis platforms.*
'''
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(readme_content)
        
        logger.info(f"Generated academic export README: {output_path}")
        return str(output_path)